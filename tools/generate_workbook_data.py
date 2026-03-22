#!/usr/bin/env python3
"""
generate_workbook_data.py
Reads the crafting spreadsheet (xlsx) and updates StratsGenerated.lua
with the latest quantities.

Usage:
    python3 tools/generate_workbook_data.py [path/to/spreadsheet.xlsx]

Defaults:
    Spreadsheet:   ~/Downloads/<latest 3-DD-YY.xlsx> or first arg
    Source Lua:    source/GoldAdvisorMidnight/Data/StratsGenerated.lua
    Workbook Lua:  source/GoldAdvisorMidnight/Data/WorkbookGenerated.lua (for item catalog)
    Output Lua:    source/GoldAdvisorMidnight/Data/StratsGenerated.lua (overwrites)
"""

import sys
import os
import re
import math
import glob as glob_mod
import json
import openpyxl
from collections import OrderedDict


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STRATS_LUA = os.path.join(REPO_ROOT, "source/GoldAdvisorMidnight/Data/StratsGenerated.lua")
WORKBOOK_LUA = os.path.join(REPO_ROOT, "source/GoldAdvisorMidnight/Data/WorkbookGenerated.lua")
MANUAL_STRATS_JSON = os.path.join(REPO_ROOT, "tools/manual_strats.json")

PATCH_TAG = "midnight-1"

# Formula profile defaults per profession (from WorkbookGenerated.lua)
FORMULA_PROFILES = {
    "alchemy":        {"multi": 30.0, "res": 15.0},
    "insc_milling":   {"multi": None, "res": 30.1},
    "insc_ink":       {"multi": 25.9, "res": 16.1},
    "jc_prospect":    {"multi": None, "res": 33.0},
    "jc_crush":       {"multi": None, "res": 35.0},
    "jc_craft":       {"multi": 29.5, "res": 33.0},
    "ench_shatter":   {"multi": None, "res": 7.8},
    "ench_craft":     {"multi": 24.5, "res": 7.8},
    "tailoring":      {"multi": 21.4, "res": 12.1},
    "blacksmithing":  {"multi": 27.9, "res": 18.7},
    "leatherworking": {"multi": 28.2, "res": 14.9},
    "engineering":    {"multi": None, "res": 36.0},
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def col_letter_to_num(letters):
    """Convert column letter(s) to 1-based index. 'A'->1, 'B'->2, 'AA'->27, etc."""
    num = 0
    for ch in letters.upper():
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num

def parse_source_block(sb):
    """Parse 'C4' -> (col_num, row_num). e.g. 'C4' -> (3, 4), 'AA10' -> (27, 10)."""
    m = re.match(r'^([A-Z]+)(\d+)$', sb.strip().upper())
    if not m:
        raise ValueError(f"Cannot parse sourceBlock: {sb!r}")
    return col_letter_to_num(m.group(1)), int(m.group(2))

def num_to_col_letter(n):
    """1-based col index to letter(s). 1->'A', 26->'Z', 27->'AA'."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(ord('A') + rem) + result
    return result

def fval(v):
    """Return float if v is numeric, else None."""
    if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
        return float(v)
    return None

def is_price_string(v):
    """True if v looks like an AH price string ('41.33', '#N/A', etc.) rather than a qty."""
    if isinstance(v, str):
        return True
    return False


# ---------------------------------------------------------------------------
# Spreadsheet readers
# ---------------------------------------------------------------------------

def read_crafts_and_ingredients(ws, crafts_row, data_col, num_reagents,
                                 max_scan_rows=12, ingredient_start_offset=1):
    """
    Read crafts count (at crafts_row/data_col) and ingredient totals
    (starting at crafts_row+ingredient_start_offset, reading num_reagents rows).
    Returns (crafts, [ingredient_total, ...]).
    """
    crafts = fval(ws.cell(crafts_row, data_col).value)
    ingredients = []
    for i in range(num_reagents):
        r = crafts_row + ingredient_start_offset + i
        v = fval(ws.cell(r, data_col).value)
        ingredients.append(v)
    return crafts, ingredients

_SKIP_LABELS = ('cost', 'profit', 'price', 'crafted price', 'per lens', 'per vial')

def scan_expected_output(ws, start_row, label_col, data_col, num_outputs=1, max_rows=25):
    """
    Scan downward from start_row looking for 'Expected...' labels in label_col.
    Excludes rows labeled as cost/profit/price. Returns list of floats found.
    """
    results = []
    for r in range(start_row, start_row + max_rows):
        lbl = ws.cell(r, label_col).value
        val = fval(ws.cell(r, data_col).value)
        if lbl and isinstance(lbl, str) and 'expected' in lbl.lower() and val is not None:
            if not any(x in lbl.lower() for x in _SKIP_LABELS):
                results.append(val)
                if len(results) >= num_outputs:
                    break
    return results


def scan_first_float_in_col(ws, start_row, label_col, data_col, max_rows=15):
    """
    Scan downward from start_row and return [first float found in data_col]
    that is not on a row labeled as cost/profit/price/crafts/starting.
    Used for JC craft strategies where the output has no 'Expected' label.
    """
    for r in range(start_row, start_row + max_rows):
        lbl = ws.cell(r, label_col).value
        val = fval(ws.cell(r, data_col).value)
        if val is not None:
            lbl_str = (lbl or "").lower()
            if not any(x in lbl_str for x in _SKIP_LABELS + ('crafts', 'starting')):
                return [val]
    return []


# ---------------------------------------------------------------------------
# Per-profession parsers
# ---------------------------------------------------------------------------

def parse_standard_block(ws, sb_col, sb_row, num_reagents, num_outputs=1):
    """
    Standard convention: sb_col/sb_row = first ingredient row, data col.
    Crafts at sb_row-1, ingredients at sb_row..sb_row+num_reagents-1.
    Returns (crafts, ingredient_totals, expected_output_totals).
    """
    label_col = sb_col - 1
    data_col = sb_col

    crafts_row = sb_row - 1
    crafts = fval(ws.cell(crafts_row, data_col).value)

    ingredients = []
    for i in range(num_reagents):
        v = fval(ws.cell(sb_row + i, data_col).value)
        ingredients.append(v)

    expected = scan_expected_output(ws, sb_row + num_reagents, label_col, data_col, num_outputs)
    return crafts, ingredients, expected


def parse_jc_prospecting_block(ws, sb_col, sb_row, num_outputs):
    """
    JC regular prospecting: sourceBlock = crafts row.
    Crafts at (sb_col, sb_row). Single ingredient (ore) at sb_row+1.
    Outputs at sb_row+3 .. sb_row+2+num_outputs.
    """
    data_col = sb_col
    crafts = fval(ws.cell(sb_row, data_col).value)
    ore_total = fval(ws.cell(sb_row + 1, data_col).value)

    # Outputs start after "Expected Output" header
    outputs = []
    for i in range(num_outputs):
        v = fval(ws.cell(sb_row + 3 + i, data_col).value)
        outputs.append(v)

    return crafts, [ore_total], outputs


def parse_jc_dazzling_block(ws, sb_col, sb_row, num_outputs):
    """
    Dazzling Thorium Prospecting: no Crafts row, ore at sb_row, outputs at sb_row+2..
    """
    data_col = sb_col
    crafts = None  # fixed yield, no crafts concept
    ore_total = fval(ws.cell(sb_row, data_col).value)

    outputs = []
    for i in range(num_outputs):
        v = fval(ws.cell(sb_row + 2 + i, data_col).value)
        outputs.append(v)

    return crafts, [ore_total], outputs


def parse_blacksmithing_block(ws, sb_col, sb_row):
    """
    Blacksmithing: sourceBlock = label col, title row.
    Data col = sb_col + 1.
    Q1 block at sb_row (rows sb_row to sb_row+17).
    Q2 block at sb_row+18 (rows sb_row+18 to sb_row+35).

    Returns:
        q1: (crafts, ingredient_totals, expected_qty)
        q2: (crafts, {Q1_col_totals: [...], Q2_col_totals: [...]}, expected_qty)
    """
    data_col = sb_col + 1  # label col B -> data col C, label F -> data G, etc.
    q2_col_q1 = data_col      # Q1 ore quantities for Q2 block
    q2_col_q2 = data_col + 1  # Q2 ore quantities for Q2 block

    # --- Q1 block ---
    q1_crafts_row = sb_row + 1
    q1_crafts = fval(ws.cell(q1_crafts_row, data_col).value)

    # Ingredients: read from crafts_row+1 while cells are numeric (not strings)
    q1_ingredients = []
    for r_off in range(2, 12):
        r = sb_row + r_off
        v = fval(ws.cell(r, data_col).value)
        lbl = ws.cell(r, sb_col).value
        if v is None or is_price_string(ws.cell(r, data_col).value):
            break
        if lbl and isinstance(lbl, str) and 'price' in lbl.lower():
            break
        q1_ingredients.append(v)

    # Expected output for Q1 block
    q1_expected_rows = scan_expected_output(ws, sb_row + 2, sb_col, data_col, num_outputs=1)
    q1_expected = q1_expected_rows[0] if q1_expected_rows else None

    # --- Q2 block ---
    q2_offset = 18  # Q2 title row = sb_row + 18
    q2_crafts_row = sb_row + q2_offset + 1
    q2_crafts = fval(ws.cell(q2_crafts_row, data_col).value)

    # Q2 ingredients: starting 3 rows after Q2 crafts (after Q1/Q2 header row)
    q2_ingredients_q1_col = []  # Q1-ore column quantities
    q2_ingredients_q2_col = []  # Q2-ore column quantities
    for r_off in range(3, 12):
        r = q2_crafts_row + r_off
        v_q1 = fval(ws.cell(r, q2_col_q1).value)
        v_q2 = fval(ws.cell(r, q2_col_q2).value)
        lbl = ws.cell(r, sb_col).value
        if v_q1 is None and v_q2 is None:
            break
        if isinstance(ws.cell(r, data_col).value, str):
            break  # hit price column
        q2_ingredients_q1_col.append(v_q1)
        q2_ingredients_q2_col.append(v_q2)

    # Expected output for Q2 block: scan from q2_crafts_row+3 onward
    q2_expected_rows = scan_expected_output(ws, q2_crafts_row + 3, sb_col, data_col, num_outputs=1)
    q2_expected = q2_expected_rows[0] if q2_expected_rows else None

    return (
        (q1_crafts, q1_ingredients, q1_expected),
        (q2_crafts, q2_ingredients_q1_col, q2_ingredients_q2_col, q2_expected),
    )


def parse_engineering_block(ws, sb_col, sb_row, num_reagents, num_outputs=1):
    """
    Engineering: no Crafts row.
    sourceBlock = data col, starting amount row.
    Returns (starting_amount, ingredient_totals, expected_qty).
    """
    data_col = sb_col
    label_col = sb_col - 1
    starting_amount = fval(ws.cell(sb_row, data_col).value)

    # Single ingredient at sb_row (the starting amount IS the ingredient total)
    ingredients = [starting_amount] + [
        fval(ws.cell(sb_row + i + 1, data_col).value)
        for i in range(num_reagents - 1)
    ]

    expected = scan_expected_output(ws, sb_row + 1, label_col, data_col, num_outputs)
    return starting_amount, ingredients, expected


def parse_jc_crushing_block(ws, sb_col, sb_row, num_reagents):
    """
    JC Crushing: 2 ingredients side-by-side in same row (sb_col and sb_col+1).
    Crafts at sb_row-1, col sb_col.
    """
    data_col = sb_col
    label_col = sb_col - 1
    crafts = fval(ws.cell(sb_row - 1, data_col).value)
    # Two ingredients in the SAME row but different columns
    ingredients = [
        fval(ws.cell(sb_row, data_col + i).value)
        for i in range(num_reagents)
    ]
    expected = scan_expected_output(ws, sb_row + 1, label_col, data_col, num_outputs=1)
    return crafts, ingredients, expected


def parse_jc_craft_sidebyside_block(ws, sb_col, sb_row, num_reagents):
    """
    JC craft strategies with ingredients side-by-side at sourceBlock row
    (e.g. Sin'dorei Lens Crafting, Sunglass Vial Crafting).
    Layout: crafts at (sb_row-1, sb_col), ingredients at (sb_row, sb_col+0..+N-1).
    Output labeled with item name (no 'Expected' label) — scanned as first float.
    """
    label_col = sb_col - 1
    data_col = sb_col
    crafts = fval(ws.cell(sb_row - 1, data_col).value)
    ingredients = [
        fval(ws.cell(sb_row, sb_col + i).value)
        for i in range(num_reagents)
    ]
    expected = scan_first_float_in_col(ws, sb_row + 1, label_col, data_col)
    return crafts, ingredients, expected


def parse_lw_scale_woven_block(ws, sb_col, sb_row, num_reagents):
    """
    Scale Woven Hide: sourceBlock = crafts row.
    Crafts at (sb_col, sb_row). Ingredient LABELS at row sb_row+1,
    ingredient QUANTITIES at row sb_row+2 (two cols side-by-side: sb_col, sb_col+1).
    Expected output: scan from sb_row+3.
    """
    data_col = sb_col
    label_col = sb_col - 1
    crafts = fval(ws.cell(sb_row, data_col).value)
    # Two ingredients side-by-side at row sb_row+2
    ingredients = [
        fval(ws.cell(sb_row + 2, sb_col + i).value)
        for i in range(num_reagents)
    ]
    expected = scan_expected_output(ws, sb_row + 3, label_col, data_col, num_outputs=1)
    return crafts, ingredients, expected


# ---------------------------------------------------------------------------
# Lua parser
# ---------------------------------------------------------------------------

def parse_strats_lua(path):
    """
    Parse StratsGenerated.lua and return a list of dicts, one per strategy.
    Each dict has the raw Lua text plus parsed key fields.
    """
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    # Split header from strategy blocks
    parts = re.split(r'(GAM_RECIPES_GENERATED\[#GAM_RECIPES_GENERATED\+1\] = )', content)
    header = parts[0]
    strategies = []
    for i in range(1, len(parts), 2):
        prefix = parts[i]
        body = parts[i + 1] if i + 1 < len(parts) else ""
        # Find the matching closing brace
        block = extract_top_level_table(body)
        strategies.append({
            "raw": prefix + block,
            "body": block,
            "parsed": parse_strategy_block(block),
        })

    return header, strategies


def extract_top_level_table(text):
    """Extract the top-level {...} block from text."""
    depth = 0
    start = text.find('{')
    if start == -1:
        return text
    end = start
    for i in range(start, len(text)):
        if text[i] == '{':
            depth += 1
        elif text[i] == '}':
            depth -= 1
            if depth == 0:
                end = i
                break
    # Include trailing whitespace/newlines
    trail = ""
    j = end + 1
    while j < len(text) and text[j] in ('\n', '\r'):
        trail += text[j]
        j += 1
    return text[start:end+1] + trail


def parse_strategy_block(block):
    """Extract key fields from a strategy block text."""
    def get_str(key):
        m = re.search(rf'{key}\s*=\s*"([^"]*)"', block)
        return m.group(1) if m else None

    def get_num(key):
        m = re.search(rf'{key}\s*=\s*([0-9.]+)', block)
        return float(m.group(1)) if m else None

    return {
        "id": get_str("id"),
        "stratName": get_str("stratName"),
        "sourceTab": get_str("sourceTab"),
        "sourceBlock": get_str("sourceBlock"),
        "profession": get_str("profession"),
        "calcMode": get_str("calcMode"),
        "formulaProfile": get_str("formulaProfile"),
        "outputQualityMode": get_str("outputQualityMode"),
        "defaultCrafts": get_num("defaultCrafts"),
        "defaultStartingAmount": get_num("defaultStartingAmount"),
        "has_rankVariants": "rankVariants" in block,
    }


# ---------------------------------------------------------------------------
# Quantity updater: Lua text-level replacements
# ---------------------------------------------------------------------------

def fmt_f(v, decimals=6):
    """Format a float to Lua float literal with given decimal places."""
    return f"{v:.{decimals}f}"


def update_num_in_block(block, key, new_val, occurrences=1):
    """Replace key = <number> in block (first `occurrences` times)."""
    pattern = rf'({re.escape(key)}\s*=\s*)[0-9.]+(\s*[,\n])'
    count = 0

    def replacer(m):
        nonlocal count
        count += 1
        if count <= occurrences:
            return m.group(1) + fmt_f(new_val) + m.group(2)
        return m.group(0)

    return re.sub(pattern, replacer, block)


def update_workbook_totals_in_block(block, totals):
    """
    Update workbookTotalQty values in reagents section.
    `totals` is a list of new values in order of reagent appearance.
    """
    pattern = r'(workbookTotalQty\s*=\s*)[0-9.]+'
    matches = list(re.finditer(pattern, block))
    if len(matches) != len(totals):
        # Mismatch - try to update as many as we have
        pass

    result = list(block)
    offset = 0
    for i, (m, t) in enumerate(zip(matches, totals)):
        if t is None:
            continue
        new_str = m.group(1) + fmt_f(t)
        start = m.start() + offset
        end = m.end() + offset
        result[start:end] = list(new_str)
        offset += len(new_str) - (m.end() - m.start())

    return "".join(result)


def update_workbook_expected_in_block(block, expected_vals):
    """
    Update workbookExpectedQty values in outputs section.
    `expected_vals` is a list in order of output appearance.
    """
    pattern = r'(workbookExpectedQty\s*=\s*)[0-9.]+'
    matches = list(re.finditer(pattern, block))

    result = list(block)
    offset = 0
    for i, (m, e) in enumerate(zip(matches, expected_vals)):
        if e is None:
            continue
        new_str = m.group(1) + fmt_f(e)
        start = m.start() + offset
        end = m.end() + offset
        result[start:end] = list(new_str)
        offset += len(new_str) - (m.end() - m.start())

    return "".join(result)


def update_qty_per_craft_in_block(block, qty_per_crafts):
    """Update qtyPerCraft values in reagents section."""
    pattern = r'(qtyPerCraft\s*=\s*)[0-9.]+'
    matches = list(re.finditer(pattern, block))

    result = list(block)
    offset = 0
    for i, (m, q) in enumerate(zip(matches, qty_per_crafts)):
        if q is None:
            continue
        new_str = m.group(1) + fmt_f(q)
        start = m.start() + offset
        end = m.end() + offset
        result[start:end] = list(new_str)
        offset += len(new_str) - (m.end() - m.start())

    return "".join(result)


def update_base_yield_in_block(block, base_yields):
    """Update baseYieldPerCraft values in outputs section."""
    pattern = r'(baseYieldPerCraft\s*=\s*)[0-9.]+'
    matches = list(re.finditer(pattern, block))

    result = list(block)
    offset = 0
    for i, (m, y) in enumerate(zip(matches, base_yields)):
        if y is None:
            continue
        new_str = m.group(1) + fmt_f(y)
        start = m.start() + offset
        end = m.end() + offset
        result[start:end] = list(new_str)
        offset += len(new_str) - (m.end() - m.start())

    return "".join(result)


# ---------------------------------------------------------------------------
# Main: process each strategy
# ---------------------------------------------------------------------------

def compute_expected_qty(sheet_expected, sheet_crafts, target_crafts, is_fixed_yield=False):
    """Scale sheet expected qty to target_crafts batch size."""
    if sheet_expected is None or sheet_crafts is None or sheet_crafts == 0:
        return None
    if is_fixed_yield:
        # For fixed calcMode, the expected qty already accounts for the batch
        # Scale to target_crafts
        return sheet_expected * (target_crafts / sheet_crafts)
    # For formula calcMode, already scaled in the sheet to sheet_crafts batch
    return sheet_expected * (target_crafts / sheet_crafts)


def process_standard_strategy(ws, strat, wb_sheet):
    """Process a strategy with the standard "Crafts:" pattern."""
    parsed = strat["parsed"]
    block = strat["body"]
    sb_col, sb_row = parse_source_block(parsed["sourceBlock"])
    profession = parsed["profession"]
    strat_name = parsed["stratName"]

    # Count reagents and outputs in the existing block
    num_reagents = len(re.findall(r'qtyPerCraft\s*=', block))
    num_outputs = len(re.findall(r'workbookExpectedQty\s*=', block))

    # Only count top-level (not in rankVariants) for some counts
    # For rankVariants strategies, outputs/reagents appear multiple times - count top-level only
    if "rankVariants" in block:
        # Count only in top-level outputs/reagents (after rankVariants block)
        pass

    is_fixed = parsed.get("calcMode") == "fixed"
    target_crafts = parsed["defaultCrafts"]
    target_starting = parsed["defaultStartingAmount"]

    # Special routing by profession/strategy
    if profession == "Blacksmithing":
        return process_blacksmithing(ws, strat, wb_sheet)

    elif strat_name == "Dazzling Thorium Prospecting":
        return process_dazzling_thorium(ws, strat, wb_sheet)

    elif profession == "Jewelcrafting" and "Prospecting" in strat_name:
        crafts, ingr, outputs = parse_jc_prospecting_block(ws, sb_col, sb_row, num_outputs)

    elif strat_name == "Crushing":
        crafts, ingr, outputs_list = parse_jc_crushing_block(ws, sb_col, sb_row, num_reagents)
        expected = scan_expected_output(ws, sb_row + 2, sb_col - 1, sb_col, 1)
        outputs = expected

    elif strat_name in ("Sin'dorei Lens Crafting", "Sunglass Vial Crafting"):
        crafts, ingr, outputs = parse_jc_craft_sidebyside_block(ws, sb_col, sb_row, num_reagents)

    elif strat_name == "Scale Woven Hide":
        crafts, ingr, outputs = parse_lw_scale_woven_block(ws, sb_col, sb_row, num_reagents)

    elif profession == "Engineering":
        starting_amount, ingr, exp = parse_engineering_block(ws, sb_col, sb_row, num_reagents, num_outputs)
        crafts = starting_amount  # Engineering: crafts == starting_amount
        outputs = exp

    else:
        crafts, ingr, outputs = parse_standard_block(ws, sb_col, sb_row, num_reagents, num_outputs)

    if crafts is None:
        print(f"  [SKIP] {strat_name}: could not read crafts count")
        return strat["raw"]

    # Compute derived values
    workbook_totals = []
    qty_per_crafts = []
    for total in ingr:
        if total is not None and crafts:
            workbook_totals.append(total)
            qty_per_crafts.append(total / crafts)
        else:
            workbook_totals.append(None)
            qty_per_crafts.append(None)

    expected_qtys = []
    for exp in outputs:
        if exp is not None:
            scaled = compute_expected_qty(exp, crafts, target_crafts, is_fixed)
            expected_qtys.append(scaled)
        else:
            expected_qtys.append(None)

    # Apply updates to block text
    new_block = block
    new_block = update_num_in_block(new_block, "defaultCrafts", crafts)
    if profession == "Engineering":
        new_block = update_num_in_block(new_block, "defaultStartingAmount", crafts)
    new_block = update_workbook_totals_in_block(new_block, workbook_totals)
    new_block = update_qty_per_craft_in_block(new_block, qty_per_crafts)
    if expected_qtys:
        new_block = update_workbook_expected_in_block(new_block, expected_qtys)
        # Also update baseYieldPerCraft for prospecting outputs
        if profession == "Jewelcrafting" and "Prospecting" in strat_name:
            base_yields = [
                (e / crafts) if (e is not None and crafts) else None
                for e in outputs
            ]
            new_block = update_base_yield_in_block(new_block, base_yields)

    prefix = strat["raw"][: strat["raw"].index(block)]
    return prefix + new_block


def process_blacksmithing(ws, strat, wb_sheet):
    """Handle Blacksmithing strategies with Q1 and Q2 rank variants."""
    parsed = strat["parsed"]
    block = strat["body"]
    strat_name = parsed["stratName"]
    sb_col, sb_row = parse_source_block(parsed["sourceBlock"])

    try:
        q1, q2 = parse_blacksmithing_block(ws, sb_col, sb_row)
    except Exception as e:
        print(f"  [SKIP] {strat_name}: error reading BS block: {e}")
        return strat["raw"]

    q1_crafts, q1_ingr, q1_expected = q1
    q2_crafts, q2_ingr_q1, q2_ingr_q2, q2_expected = q2

    if q1_crafts is None or q2_crafts is None:
        print(f"  [SKIP] {strat_name}: could not read Q1/Q2 crafts")
        return strat["raw"]

    target_crafts = q2_crafts  # defaultCrafts comes from Q2 block

    # For rankVariants.lowest: ingredient totals scaled to target_crafts
    lowest_totals = [
        (v / q1_crafts * target_crafts) if v is not None else None
        for v in q1_ingr
    ]
    lowest_qty_per_crafts = [
        (v / q1_crafts) if v is not None else None
        for v in q1_ingr
    ]
    lowest_expected = compute_expected_qty(q1_expected, q1_crafts, target_crafts)

    # For rankVariants.highest: ingredient totals from Q2 block
    # Q2 has: some Q1-col quantities + some Q2-col quantities
    # The Q2 block stores Q1+Q2 ore mixed
    # Build highest totals: use Q1-col values (Q2-col values are additional Q2 ore)
    highest_totals = []
    highest_qty_per_crafts = []
    for i, (v1, v2) in enumerate(zip(q2_ingr_q1, q2_ingr_q2)):
        if v2 is not None:
            # This ingredient has a Q2 variant (separate Q2 ore)
            highest_totals.append(v1)  # Q1 ore total
            highest_qty_per_crafts.append((v1 / q2_crafts) if v1 else None)
            highest_totals.append(v2)  # Q2 ore total
            highest_qty_per_crafts.append((v2 / q2_crafts) if v2 else None)
        else:
            # Single column (flux or ingot - not split into Q1/Q2)
            highest_totals.append(v1)
            highest_qty_per_crafts.append((v1 / q2_crafts) if v1 else None)

    highest_expected = compute_expected_qty(q2_expected, q2_crafts, target_crafts)

    # Apply updates - this is complex because rankVariants has multiple qty tables
    new_block = block
    new_block = update_num_in_block(new_block, "defaultCrafts", target_crafts)
    new_block = update_num_in_block(new_block, "defaultStartingAmount", parsed["defaultStartingAmount"])

    # Update workbookExpectedQty (multiple occurrences: lowest, highest, top-level)
    all_expected = [lowest_expected, highest_expected, lowest_expected]  # lowest, highest, top-level = same as lowest
    new_block = update_workbook_expected_in_block(new_block, all_expected)

    # Update workbookTotalQty: lowest reagents, highest reagents, top-level reagents
    all_totals = lowest_totals + highest_totals + lowest_totals
    new_block = update_workbook_totals_in_block(new_block, all_totals)

    # Update qtyPerCraft: same ordering
    all_qpc = lowest_qty_per_crafts + highest_qty_per_crafts + lowest_qty_per_crafts
    new_block = update_qty_per_craft_in_block(new_block, all_qpc)

    prefix = strat["raw"][: strat["raw"].index(block)]
    return prefix + new_block


def process_dazzling_thorium(ws, strat, wb_sheet):
    """
    Dazzling Thorium Prospecting: fix to Q1-only outputs and update quantities.
    The spreadsheet says 'This only produced q1' - all outputs are Q1 items.
    """
    parsed = strat["parsed"]
    block = strat["body"]
    sb_col, sb_row = parse_source_block(parsed["sourceBlock"])

    # Count outputs from existing block
    num_outputs = len(re.findall(r'workbookExpectedQty\s*=', block))

    crafts, ingr, outputs = parse_jc_dazzling_block(ws, sb_col, sb_row, num_outputs)

    target_crafts = parsed["defaultCrafts"] or 1000.0

    # Update workbookTotalQty for the single reagent
    ore_total = ingr[0]
    if ore_total is not None:
        block = update_workbook_totals_in_block(block, [ore_total])
        ore_qpc = ore_total / target_crafts if target_crafts else 1.0
        block = update_qty_per_craft_in_block(block, [ore_qpc])

    # Update workbookExpectedQty for each output
    scaled_outputs = [
        o for o in outputs
    ]
    block = update_workbook_expected_in_block(block, scaled_outputs)

    # Update baseYieldPerCraft for outputs
    if target_crafts:
        base_yields = [(o / target_crafts) if o is not None else None for o in scaled_outputs]
        block = update_base_yield_in_block(block, base_yields)

    # Fix to Q1-only outputs: change itemIDs from { x, y } to { x } for each output
    # Only keep the FIRST item ID (Q1) for each output in the outputs section
    # We identify "Flawless" gems as having Q1/Q2 - keep only Q1 (first ID)
    def q1_only_output_ids(m):
        ids = m.group(1)
        # If has two IDs, keep only the first
        first_id = re.match(r'\s*(\d+)', ids)
        if first_id and ',' in ids:
            return m.group(0).replace(ids, f' {first_id.group(1)} ')
        return m.group(0)

    # Only fix itemIDs in the outputs section (not reagents)
    outputs_section_start = block.find('outputs = {')
    reagents_section_start = block.find('reagents = {')

    if outputs_section_start != -1 and reagents_section_start != -1:
        outputs_section = block[outputs_section_start:reagents_section_start]
        outputs_section_fixed = re.sub(
            r'itemIDs = \{([^}]+)\}',
            q1_only_output_ids,
            outputs_section
        )
        block = block[:outputs_section_start] + outputs_section_fixed + block[reagents_section_start:]

    prefix = strat["raw"][: strat["raw"].index(strat["body"])]
    return prefix + block


# ---------------------------------------------------------------------------
# Manual strat support
# ---------------------------------------------------------------------------

def render_ids(ids):
    """Render a list of item IDs as a Lua array literal."""
    return "{ " + ", ".join(str(i) for i in ids) + " }"


def render_manual_strat_lua(entry):
    """Convert a manual_strats.json entry dict to a Lua GAM_RECIPES_GENERATED block."""
    lines = []
    lines.append(f'-- MANUAL: {entry["stratName"]} (not in spreadsheet; migrate when data is available)')
    lines.append('GAM_RECIPES_GENERATED[#GAM_RECIPES_GENERATED+1] = {')
    lines.append(f'  id = "{entry["id"]}",')
    lines.append(f'  patchTag = "{entry.get("patchTag", PATCH_TAG)}",')
    lines.append(f'  profession = "{entry["profession"]}",')
    lines.append(f'  stratName = "{entry["stratName"]}",')
    lines.append(f'  sourceTab = "Manual",')
    lines.append(f'  sourceBlock = nil,')
    lines.append(f'  defaultStartingAmount = {entry.get("defaultStartingAmount", 1.0):.6f},')
    lines.append(f'  defaultCrafts = {entry.get("defaultCrafts", 1.0):.6f},')
    lines.append(f'  formulaProfile = "{entry["formulaProfile"]}",')
    lines.append(f'  calcMode = "{entry.get("calcMode", "formula")}",')
    lines.append(f'  qualityPolicy = "{entry.get("qualityPolicy", "normal")}",')
    lines.append(f'  outputQualityMode = "{entry.get("outputQualityMode", "rank_policy")}",')
    notes = entry.get("notes", "").replace('"', '\\"')
    lines.append(f'  notes = "{notes}",')
    lines.append('  outputs = {')
    for out in entry.get("outputs", []):
        lines.append('    {')
        lines.append(f'      itemRef = "{out["itemRef"]}",')
        lines.append(f'      itemIDs = {render_ids(out["itemIDs"])},')
        lines.append(f'      baseYieldPerCraft = {out.get("baseYieldPerCraft", 1.0):.6f},')
        lines.append(f'      baseYield = {out.get("baseYield", 1.0):.6f},')
        lines.append(f'      workbookExpectedQty = {out.get("workbookExpectedQty", 1.0):.6f},')
        lines.append('    },')
    lines.append('  },')
    lines.append('  reagents = {')
    for r in entry.get("reagents", []):
        lines.append('    {')
        lines.append(f'      itemRef = "{r["itemRef"]}",')
        lines.append(f'      itemIDs = {render_ids(r["itemIDs"])},')
        lines.append(f'      qtyPerCraft = {r.get("qtyPerCraft", 1.0):.6f},')
        lines.append(f'      qtyPerStart = {r.get("qtyPerStart", 1.0):.6f},')
        lines.append(f'      workbookTotalQty = {r.get("workbookTotalQty", 1.0):.6f},')
        lines.append('    },')
    lines.append('  },')
    lines.append('}')
    return "\n".join(lines) + "\n"


def load_manual_strats():
    """Load tools/manual_strats.json if it exists; return list of entries."""
    if not os.path.exists(MANUAL_STRATS_JSON):
        return []
    with open(MANUAL_STRATS_JSON, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def find_latest_spreadsheet():
    """Find the most recently downloaded spreadsheet matching the naming pattern."""
    downloads = os.path.expanduser("~/Downloads")
    candidates = sorted(
        glob_mod.glob(os.path.join(downloads, "[0-9]-[0-9][0-9]-[0-9][0-9]*.xlsx")) +
        glob_mod.glob(os.path.join(downloads, "[0-9][0-9]-[0-9][0-9]-[0-9][0-9]*.xlsx")),
        key=os.path.getmtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def main():
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = find_latest_spreadsheet()
        if not xlsx_path:
            print("ERROR: No spreadsheet found in ~/Downloads. Pass path as argument.")
            sys.exit(1)

    print(f"Spreadsheet: {xlsx_path}")
    print(f"Source Lua:  {STRATS_LUA}")
    print()

    # Load workbook
    print("Loading spreadsheet (data_only)...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Parse existing Lua
    print("Parsing existing StratsGenerated.lua...")
    header, strategies = parse_strats_lua(STRATS_LUA)

    print(f"Found {len(strategies)} strategies to process.\n")

    updated_blocks = []
    for strat in strategies:
        parsed = strat["parsed"]
        name = parsed.get("stratName", "?")
        tab = parsed.get("sourceTab", "?")
        sb = parsed.get("sourceBlock", "?")

        if tab not in wb.sheetnames:
            print(f"  [SKIP] {name}: sheet '{tab}' not found in workbook")
            updated_blocks.append(strat["raw"])
            continue

        ws = wb[tab]
        print(f"  Processing {name} ({tab}:{sb})...")
        try:
            new_raw = process_standard_strategy(ws, strat, wb)
            updated_blocks.append(new_raw)
        except Exception as e:
            import traceback
            print(f"  [ERROR] {name}: {e}")
            traceback.print_exc()
            updated_blocks.append(strat["raw"])

    # Append manual strats from tools/manual_strats.json
    manual_entries = load_manual_strats()
    manual_blocks = []
    if manual_entries:
        print(f"\nAppending {len(manual_entries)} manual strat(s) from {MANUAL_STRATS_JSON}")
        for entry in manual_entries:
            name = entry.get("stratName", entry.get("id", "?"))
            print(f"  + {name}")
            manual_blocks.append(render_manual_strat_lua(entry))

    # Write output
    print(f"\nWriting updated Lua to {STRATS_LUA}")
    output = header + "".join(updated_blocks)
    if manual_blocks:
        output = output.rstrip('\n') + '\n\n'
        output += "-- ===== MANUAL STRATS (from tools/manual_strats.json) =====\n"
        output += "\n".join(manual_blocks)
    # Ensure single newline at end
    output = output.rstrip('\n') + '\n'

    with open(STRATS_LUA, "w", encoding="utf-8") as f:
        f.write(output)

    print("Done!")
    print()
    print("Next steps:")
    print("  1. Review the diff:  git diff source/GoldAdvisorMidnight/Data/StratsGenerated.lua")
    print("  2. Test in WoW to confirm no Lua errors")
    print("  3. Commit: git add source/ && git commit -m 'chore(data): regenerate from spreadsheet'")


if __name__ == "__main__":
    main()
