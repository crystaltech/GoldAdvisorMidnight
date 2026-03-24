#!/usr/bin/env python3
"""
compare_strats.py
Reads the crafting spreadsheet and StratsGenerated.lua, then produces a
detailed diff report showing any mismatches between the two.

Checks per strategy:
  - defaultCrafts (Lua) vs crafts count read from spreadsheet
  - Each reagent's qtyPerCraft (Lua) vs (sheet_total / sheet_crafts)
  - Each output's workbookExpectedQty (Lua) vs spreadsheet expected output
  - Strategies present in Lua but not parseable from spreadsheet (or vice versa)

Usage:
    python3 tools/compare_strats.py [path/to/spreadsheet.xlsx]
"""

import sys
import os
import re
import math
import glob as glob_mod
import openpyxl
from collections import OrderedDict

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STRATS_LUA = os.path.join(REPO_ROOT, "source/GoldAdvisorMidnight/Data/StratsGenerated.lua")

TOLERANCE = 0.01  # Ignore floating-point differences smaller than this

# ---------------------------------------------------------------------------
# Helpers (copied from generate_workbook_data.py)
# ---------------------------------------------------------------------------

def col_letter_to_num(letters):
    num = 0
    for ch in letters.upper():
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num

def parse_source_block(sb):
    m = re.match(r'^([A-Z]+)(\d+)$', sb.strip().upper())
    if not m:
        raise ValueError(f"Cannot parse sourceBlock: {sb!r}")
    return col_letter_to_num(m.group(1)), int(m.group(2))

def num_to_col_letter(n):
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(ord('A') + rem) + result
    return result

def fval(v):
    if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
        return float(v)
    return None

def is_price_string(v):
    return isinstance(v, str)

_SKIP_LABELS = ('cost', 'profit', 'price', 'crafted price', 'value per', 'per lens', 'per vial')

def scan_expected_output(ws, start_row, label_col, data_col, num_outputs=1, max_rows=25):
    results = []
    for r in range(start_row, start_row + max_rows):
        lbl = ws.cell(r, label_col).value
        val = fval(ws.cell(r, data_col).value)
        if lbl and isinstance(lbl, str) and 'expected' in lbl.lower() and val is not None:
            if not any(x in lbl.lower() for x in _SKIP_LABELS):
                results.append(val)
                if len(results) >= num_outputs:
                    break
    return results

def scan_first_float_in_col(ws, start_row, label_col, data_col, max_rows=15):
    for r in range(start_row, start_row + max_rows):
        lbl = ws.cell(r, label_col).value
        val = fval(ws.cell(r, data_col).value)
        if val is not None:
            lbl_str = (lbl or "").lower()
            if not any(x in lbl_str for x in _SKIP_LABELS + ('crafts', 'starting')):
                return [val]
    return []

# ---------------------------------------------------------------------------
# Per-profession parsers (copied from generate_workbook_data.py)
# ---------------------------------------------------------------------------

def parse_standard_block(ws, sb_col, sb_row, num_reagents, num_outputs=1):
    label_col = sb_col - 1
    data_col = sb_col
    crafts_row = sb_row - 1
    crafts = fval(ws.cell(crafts_row, data_col).value)
    ingredients = []
    for i in range(num_reagents):
        v = fval(ws.cell(sb_row + i, data_col).value)
        ingredients.append(v)
    expected = scan_expected_output(ws, sb_row + num_reagents, label_col, data_col, num_outputs)
    return crafts, ingredients, expected

def parse_jc_prospecting_block(ws, sb_col, sb_row, num_outputs):
    data_col = sb_col
    crafts = fval(ws.cell(sb_row, data_col).value)
    ore_total = fval(ws.cell(sb_row + 1, data_col).value)
    outputs = []
    for i in range(num_outputs):
        v = fval(ws.cell(sb_row + 3 + i, data_col).value)
        outputs.append(v)
    return crafts, [ore_total], outputs

def parse_jc_dazzling_block(ws, sb_col, sb_row, num_outputs):
    data_col = sb_col
    crafts = None  # fixed yield, no crafts concept
    ore_total = fval(ws.cell(sb_row, data_col).value)
    outputs = []
    for i in range(num_outputs):
        v = fval(ws.cell(sb_row + 2 + i, data_col).value)
        outputs.append(v)
    return crafts, [ore_total], outputs

def parse_blacksmithing_block(ws, sb_col, sb_row):
    data_col = sb_col + 1
    q2_col_q1 = data_col
    q2_col_q2 = data_col + 1

    # Q1 block
    q1_crafts_row = sb_row + 1
    q1_crafts = fval(ws.cell(q1_crafts_row, data_col).value)
    q1_ingredients = []
    for r_off in range(2, 12):
        r = sb_row + r_off
        v = fval(ws.cell(r, data_col).value)
        lbl = ws.cell(r, sb_col).value
        if v is None or is_price_string(ws.cell(r, data_col).value):
            break
        if lbl and isinstance(lbl, str) and 'price' in lbl.lower():
            break
        q1_ingredients.append(v)

    q1_expected_rows = scan_expected_output(ws, sb_row + 2, sb_col, data_col, num_outputs=1)
    q1_expected = q1_expected_rows[0] if q1_expected_rows else None

    # Q2 block
    q2_offset = 18
    q2_crafts_row = sb_row + q2_offset + 1
    q2_crafts = fval(ws.cell(q2_crafts_row, data_col).value)
    q2_ingredients_q1_col = []
    q2_ingredients_q2_col = []
    # Same workbook layout as generate_workbook_data.py: the first Q2 reagent row
    # is two rows below the crafts row, immediately after the Q1/Q2 header row.
    for r_off in range(2, 12):
        r = q2_crafts_row + r_off
        v_q1 = fval(ws.cell(r, q2_col_q1).value)
        v_q2 = fval(ws.cell(r, q2_col_q2).value)
        lbl = ws.cell(r, sb_col).value
        if v_q1 is None and v_q2 is None:
            break
        if isinstance(ws.cell(r, data_col).value, str):
            break
        q2_ingredients_q1_col.append(v_q1)
        q2_ingredients_q2_col.append(v_q2)

    q2_expected_rows = scan_expected_output(ws, q2_crafts_row + 3, sb_col, data_col, num_outputs=1)
    q2_expected = q2_expected_rows[0] if q2_expected_rows else None

    return (
        (q1_crafts, q1_ingredients, q1_expected),
        (q2_crafts, q2_ingredients_q1_col, q2_ingredients_q2_col, q2_expected),
    )

def parse_engineering_block(ws, sb_col, sb_row, num_reagents, num_outputs=1):
    data_col = sb_col
    label_col = sb_col - 1
    starting_amount = fval(ws.cell(sb_row, data_col).value)
    ingredients = [starting_amount] + [
        fval(ws.cell(sb_row + i + 1, data_col).value)
        for i in range(num_reagents - 1)
    ]
    expected = scan_expected_output(ws, sb_row + 1, label_col, data_col, num_outputs)
    return starting_amount, ingredients, expected

def parse_jc_crushing_block(ws, sb_col, sb_row, num_reagents):
    data_col = sb_col
    label_col = sb_col - 1
    crafts = fval(ws.cell(sb_row - 1, data_col).value)
    ingredients = [
        fval(ws.cell(sb_row, data_col + i).value)
        for i in range(num_reagents)
    ]
    expected = scan_expected_output(ws, sb_row + 1, label_col, data_col, num_outputs=1)
    return crafts, ingredients, expected

def parse_jc_craft_sidebyside_block(ws, sb_col, sb_row, num_reagents):
    label_col = sb_col - 1
    data_col = sb_col
    crafts = fval(ws.cell(sb_row - 1, data_col).value)
    ingredients = [
        fval(ws.cell(sb_row, sb_col + i).value)
        for i in range(num_reagents)
    ]
    expected = scan_first_float_in_col(ws, sb_row + 1, label_col, data_col)
    return crafts, ingredients, expected

def parse_lw_scale_woven_block(ws, sb_col, sb_row, num_reagents):
    data_col = sb_col
    label_col = sb_col - 1
    crafts = fval(ws.cell(sb_row, data_col).value)
    ingredients = [
        fval(ws.cell(sb_row + 2, sb_col + i).value)
        for i in range(num_reagents)
    ]
    expected = scan_expected_output(ws, sb_row + 3, label_col, data_col, num_outputs=1)
    return crafts, ingredients, expected

# ---------------------------------------------------------------------------
# Lua parser (copied from generate_workbook_data.py)
# ---------------------------------------------------------------------------

def extract_top_level_table(text):
    depth = 0
    start = text.find('{')
    if start == -1:
        return text
    end = start
    for i in range(start, len(text)):
        if text[i] == '{':
            depth += 1
        elif text[i] == '}':
            depth -= 1
            if depth == 0:
                end = i
                break
    trail = ""
    j = end + 1
    while j < len(text) and text[j] in ('\n', '\r'):
        trail += text[j]
        j += 1
    return text[start:end+1] + trail

def parse_strategy_block(block):
    def get_str(key):
        m = re.search(rf'{key}\s*=\s*"([^"]*)"', block)
        return m.group(1) if m else None

    def get_num(key):
        m = re.search(rf'{key}\s*=\s*([0-9.]+)', block)
        return float(m.group(1)) if m else None

    return {
        "id": get_str("id"),
        "stratName": get_str("stratName"),
        "sourceTab": get_str("sourceTab"),
        "sourceBlock": get_str("sourceBlock"),
        "profession": get_str("profession"),
        "calcMode": get_str("calcMode"),
        "formulaProfile": get_str("formulaProfile"),
        "outputQualityMode": get_str("outputQualityMode"),
        "defaultCrafts": get_num("defaultCrafts"),
        "defaultStartingAmount": get_num("defaultStartingAmount"),
        "has_rankVariants": "rankVariants" in block,
    }

def parse_strats_lua(path):
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()
    parts = re.split(r'(GAM_RECIPES_GENERATED\[#GAM_RECIPES_GENERATED\+1\] = )', content)
    header = parts[0]
    strategies = []
    for i in range(1, len(parts), 2):
        prefix = parts[i]
        body = parts[i + 1] if i + 1 < len(parts) else ""
        block = extract_top_level_table(body)
        strategies.append({
            "raw": prefix + block,
            "body": block,
            "parsed": parse_strategy_block(block),
        })
    return header, strategies

# ---------------------------------------------------------------------------
# Lua deep parsers: extract qtyPerCraft and workbookExpectedQty values
# ---------------------------------------------------------------------------

def extract_qty_per_craft_values(block):
    """Return list of (index, value) for all qtyPerCraft occurrences in block."""
    return [float(m.group(1)) for m in re.finditer(r'qtyPerCraft\s*=\s*([0-9.]+)', block)]

def extract_workbook_expected_values(block):
    """Return list of values for all workbookExpectedQty occurrences in block."""
    return [float(m.group(1)) for m in re.finditer(r'workbookExpectedQty\s*=\s*([0-9.]+)', block)]

def extract_workbook_total_values(block):
    """Return list of values for all workbookTotalQty occurrences in block."""
    return [float(m.group(1)) for m in re.finditer(r'workbookTotalQty\s*=\s*([0-9.]+)', block)]

def extract_reagent_names(block):
    """Return list of itemRef names for all reagents."""
    reagents_section = re.search(r'reagents\s*=\s*\{(.*)\}\s*\}', block, re.DOTALL)
    if not reagents_section:
        return []
    return re.findall(r'itemRef\s*=\s*"([^"]*)"', reagents_section.group(1))

def extract_output_names(block):
    """Return list of itemRef names for all outputs."""
    outputs_section = re.search(r'outputs\s*=\s*\{(.*?)reagents\s*=', block, re.DOTALL)
    if not outputs_section:
        # Try without reagents following
        outputs_section = re.search(r'outputs\s*=\s*\{(.+)\}', block, re.DOTALL)
    if not outputs_section:
        return []
    return re.findall(r'itemRef\s*=\s*"([^"]*)"', outputs_section.group(1))

def extract_rankvariant_sections(block):
    """
    For Blacksmithing rankVariants blocks, extract lowest and highest sub-blocks.
    Returns (lowest_block, highest_block) or (None, None).
    """
    rv_match = re.search(r'rankVariants\s*=\s*\{', block)
    if not rv_match:
        return None, None

    # Find the rankVariants table
    rv_start = rv_match.start()
    rv_brace = block.index('{', rv_start)
    depth = 0
    rv_end = rv_brace
    for i in range(rv_brace, len(block)):
        if block[i] == '{':
            depth += 1
        elif block[i] == '}':
            depth -= 1
            if depth == 0:
                rv_end = i
                break
    rv_content = block[rv_brace:rv_end+1]

    # Extract lowest and highest sub-blocks
    lowest_match = re.search(r'lowest\s*=\s*\{', rv_content)
    highest_match = re.search(r'highest\s*=\s*\{', rv_content)

    def extract_sub_block(content, match):
        if not match:
            return None
        brace_start = content.index('{', match.start())
        depth = 0
        end = brace_start
        for i in range(brace_start, len(content)):
            if content[i] == '{':
                depth += 1
            elif content[i] == '}':
                depth -= 1
                if depth == 0:
                    end = i
                    break
        return content[brace_start:end+1]

    lowest_block = extract_sub_block(rv_content, lowest_match)
    highest_block = extract_sub_block(rv_content, highest_match)
    return lowest_block, highest_block

# ---------------------------------------------------------------------------
# Comparison logic
# ---------------------------------------------------------------------------

def close_enough(a, b, tol=TOLERANCE):
    if a is None or b is None:
        return False
    return abs(a - b) <= tol

def compare_values(name, label, lua_val, sheet_val, mismatches):
    """Compare two values and record any mismatch."""
    if lua_val is None and sheet_val is None:
        return
    if lua_val is None:
        mismatches.append(f"  {label}: Lua=None, Sheet={sheet_val:.4f}  [Lua missing]")
        return
    if sheet_val is None:
        mismatches.append(f"  {label}: Lua={lua_val:.4f}, Sheet=None  [Sheet missing]")
        return
    if not close_enough(lua_val, sheet_val):
        diff = lua_val - sheet_val
        mismatches.append(
            f"  {label}: Lua={lua_val:.6f}, Sheet={sheet_val:.6f}  (diff={diff:+.6f})"
        )

def compare_standard_strategy(ws, strat, wb):
    """
    Compare a standard (non-BS) strategy.
    Returns list of mismatch strings.
    """
    parsed = strat["parsed"]
    block = strat["body"]
    strat_name = parsed["stratName"]
    profession = parsed["profession"]
    sb_col, sb_row = parse_source_block(parsed["sourceBlock"])

    lua_crafts = parsed["defaultCrafts"]
    lua_starting = parsed["defaultStartingAmount"]
    lua_qpc = extract_qty_per_craft_values(block)
    lua_expected = extract_workbook_expected_values(block)

    num_reagents = len(lua_qpc)
    num_outputs = len(lua_expected)

    mismatches = []

    # Route by special strategy type
    if profession == "Blacksmithing":
        return compare_blacksmithing(ws, strat)

    elif strat_name == "Dazzling Thorium Prospecting":
        # No defaultCrafts comparison; ore qtyPerCraft = ore_total / target_crafts
        crafts, ingr, outputs = parse_jc_dazzling_block(ws, sb_col, sb_row, num_outputs)
        target_crafts = lua_crafts or 1000.0
        if ingr[0] is not None and target_crafts:
            sheet_qpc = ingr[0] / target_crafts
            if lua_qpc:
                compare_values(strat_name, "reagent[0] qtyPerCraft", lua_qpc[0], sheet_qpc, mismatches)
        # Expected outputs: sheet values are already the expected quantities (no scaling)
        for i, (lua_e, sheet_e) in enumerate(zip(lua_expected, outputs)):
            compare_values(strat_name, f"output[{i}] workbookExpectedQty", lua_e, sheet_e, mismatches)
        return mismatches

    elif profession == "Jewelcrafting" and "Prospecting" in strat_name:
        crafts, ingr, outputs = parse_jc_prospecting_block(ws, sb_col, sb_row, num_outputs)

    elif strat_name == "Crushing":
        crafts, ingr, _ = parse_jc_crushing_block(ws, sb_col, sb_row, num_reagents)
        expected = scan_expected_output(ws, sb_row + 2, sb_col - 1, sb_col, 1)
        outputs = expected

    elif strat_name in ("Sin'dorei Lens Crafting", "Sunglass Vial Crafting"):
        crafts, ingr, outputs = parse_jc_craft_sidebyside_block(ws, sb_col, sb_row, num_reagents)

    elif strat_name == "Scale Woven Hide":
        crafts, ingr, outputs = parse_lw_scale_woven_block(ws, sb_col, sb_row, num_reagents)

    elif profession == "Engineering":
        starting_amount, ingr, exp = parse_engineering_block(ws, sb_col, sb_row, num_reagents, num_outputs)
        crafts = starting_amount
        outputs = exp
        # Engineering: compare defaultStartingAmount instead of defaultCrafts
        compare_values(strat_name, "defaultStartingAmount", lua_starting, starting_amount, mismatches)

    else:
        crafts, ingr, outputs = parse_standard_block(ws, sb_col, sb_row, num_reagents, num_outputs)

    if crafts is None:
        mismatches.append(f"  [SKIP] Could not read crafts count from spreadsheet")
        return mismatches

    # Compare defaultCrafts (skip for Engineering — already done above)
    if profession != "Engineering":
        compare_values(strat_name, "defaultCrafts", lua_crafts, crafts, mismatches)

    # Compare qtyPerCraft for each reagent
    for i, (sheet_total, lua_q) in enumerate(zip(ingr, lua_qpc)):
        if sheet_total is not None and crafts:
            sheet_qpc = sheet_total / crafts
            compare_values(strat_name, f"reagent[{i}] qtyPerCraft", lua_q, sheet_qpc, mismatches)

    # Compare workbookExpectedQty for each output
    # Expected in sheet is at sheet's crafts level; Lua stores at lua_crafts level
    # Scale: sheet_expected * (lua_crafts / crafts)
    if lua_crafts and crafts:
        scale = lua_crafts / crafts
    else:
        scale = 1.0

    for i, (lua_e, sheet_e) in enumerate(zip(lua_expected, outputs)):
        if sheet_e is not None:
            scaled_sheet_e = sheet_e * scale
            compare_values(strat_name, f"output[{i}] workbookExpectedQty", lua_e, scaled_sheet_e, mismatches)

    return mismatches


def compare_blacksmithing(ws, strat):
    """Compare a Blacksmithing strategy with rankVariants."""
    parsed = strat["parsed"]
    block = strat["body"]
    strat_name = parsed["stratName"]
    sb_col, sb_row = parse_source_block(parsed["sourceBlock"])

    mismatches = []

    try:
        q1, q2 = parse_blacksmithing_block(ws, sb_col, sb_row)
    except Exception as e:
        mismatches.append(f"  [ERROR] Could not parse BS block: {e}")
        return mismatches

    q1_crafts, q1_ingr, q1_expected = q1
    q2_crafts, q2_ingr_q1, q2_ingr_q2, q2_expected = q2

    if q1_crafts is None or q2_crafts is None:
        mismatches.append(f"  [SKIP] Could not read Q1/Q2 crafts")
        return mismatches

    target_crafts = q2_crafts  # defaultCrafts comes from Q2 block
    lua_crafts = parsed["defaultCrafts"]
    compare_values(strat_name, "defaultCrafts", lua_crafts, target_crafts, mismatches)

    # Extract rankVariant sub-blocks
    lowest_block, highest_block = extract_rankvariant_sections(block)

    # --- Compare lowest (Q1) variant ---
    if lowest_block:
        lowest_qpc = extract_qty_per_craft_values(lowest_block)
        lowest_expected = extract_workbook_expected_values(lowest_block)

        for i, (sheet_total, lua_q) in enumerate(zip(q1_ingr, lowest_qpc)):
            if sheet_total is not None and q1_crafts:
                sheet_qpc = sheet_total / q1_crafts
                compare_values(strat_name, f"lowest/reagent[{i}] qtyPerCraft", lua_q, sheet_qpc, mismatches)

        if lowest_expected and q1_expected is not None:
            scaled = q1_expected * (target_crafts / q1_crafts) if q1_crafts else q1_expected
            compare_values(strat_name, "lowest workbookExpectedQty", lowest_expected[0], scaled, mismatches)
    else:
        mismatches.append(f"  [WARN] No rankVariants.lowest block found in Lua")

    # --- Compare highest (Q2) variant ---
    if highest_block:
        highest_qpc = extract_qty_per_craft_values(highest_block)
        highest_expected = extract_workbook_expected_values(highest_block)

        # Reconstruct the expected qtyPerCraft list for highest (mirrors process_blacksmithing)
        expected_highest_qpc = []
        for i, (v1, v2) in enumerate(zip(q2_ingr_q1, q2_ingr_q2)):
            if v2 is not None:
                expected_highest_qpc.append((v1 / q2_crafts) if v1 else None)
                expected_highest_qpc.append((v2 / q2_crafts) if v2 else None)
            else:
                expected_highest_qpc.append((v1 / q2_crafts) if v1 else None)

        for i, (lua_q, sheet_qpc) in enumerate(zip(highest_qpc, expected_highest_qpc)):
            if sheet_qpc is not None:
                compare_values(strat_name, f"highest/reagent[{i}] qtyPerCraft", lua_q, sheet_qpc, mismatches)

        if highest_expected and q2_expected is not None:
            scaled = q2_expected  # Q2 crafts == target_crafts, so scale is 1.0
            compare_values(strat_name, "highest workbookExpectedQty", highest_expected[0], scaled, mismatches)
    else:
        mismatches.append(f"  [WARN] No rankVariants.highest block found in Lua")

    return mismatches


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def find_latest_spreadsheet():
    downloads = os.path.expanduser("~/Downloads")
    candidates = sorted(
        glob_mod.glob(os.path.join(downloads, "[0-9]-[0-9][0-9]-[0-9][0-9]*.xlsx")) +
        glob_mod.glob(os.path.join(downloads, "[0-9][0-9]-[0-9][0-9]-[0-9][0-9]*.xlsx")),
        key=os.path.getmtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def main():
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = find_latest_spreadsheet()
        if not xlsx_path:
            print("ERROR: No spreadsheet found in ~/Downloads. Pass path as argument.")
            sys.exit(1)

    print("=" * 70)
    print("STRATS COMPARISON REPORT")
    print("=" * 70)
    print(f"Spreadsheet: {xlsx_path}")
    print(f"Lua file:    {STRATS_LUA}")
    print(f"Tolerance:   {TOLERANCE}")
    print()

    print("Loading spreadsheet (data_only)...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print("Parsing StratsGenerated.lua...")
    header, strategies = parse_strats_lua(STRATS_LUA)
    print(f"Found {len(strategies)} strategies in Lua.\n")

    total_mismatches = 0
    skipped = []
    ok_count = 0

    all_results = []

    for strat in strategies:
        parsed = strat["parsed"]
        name = parsed.get("stratName", "?")
        tab = parsed.get("sourceTab", "?")
        sb = parsed.get("sourceBlock", "?")
        profession = parsed.get("profession", "?")

        if tab not in wb.sheetnames:
            skipped.append(f"  {name}  (sheet '{tab}' not in workbook)")
            continue

        ws = wb[tab]

        try:
            mismatches = compare_standard_strategy(ws, strat, wb)
        except Exception as e:
            import traceback
            mismatches = [f"  [EXCEPTION] {e}\n{traceback.format_exc()}"]

        if mismatches:
            total_mismatches += len(mismatches)
            all_results.append((name, tab, sb, profession, mismatches))
        else:
            ok_count += 1

    # Print OK strategies summary
    print(f"Strategies with NO mismatches: {ok_count}")
    print()

    # Print skipped
    if skipped:
        print(f"SKIPPED ({len(skipped)}) — sheet not found in workbook:")
        for s in skipped:
            print(s)
        print()

    # Print mismatches
    if all_results:
        print(f"MISMATCHES FOUND in {len(all_results)} strategies ({total_mismatches} total issues):")
        print("-" * 70)
        for name, tab, sb, profession, mismatches in all_results:
            print(f"\n[{profession}] {name}  ({tab}:{sb})")
            for m in mismatches:
                print(m)
    else:
        print("No mismatches found! Lua file is consistent with the spreadsheet.")

    print()
    print("=" * 70)
    print(f"SUMMARY: {ok_count} OK, {len(all_results)} with mismatches, {len(skipped)} skipped")
    print("=" * 70)


if __name__ == "__main__":
    main()
